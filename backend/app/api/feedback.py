"""
Feedback Analysis API

Endpoints:
    POST /upload-feedback              — upload a structured feedback file (.json or .csv)
    POST /analyze-feedback             — submit raw feedback text (kept for future LLM use)
    GET  /feedback-datasets/{project_id}
    GET  /feedback-dataset/{dataset_id}
    DELETE /feedback-dataset/{dataset_id}
    GET  /export-dataset/{dataset_id}  — export as .xlsx
    GET  /export-dataset/{dataset_id}/csv — export as Sensecap-compatible CSV

Pipeline order for file upload:
    1. Parse structured file   — extract FeedbackSegment list from JSON or CSV
    2. Save to SQLite          — persist segments BEFORE Agent 2 runs
    3. Video Optimization Agent — segments + metadata → recommendations + editing plan
"""

import json
import csv
import io
import re
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.schemas.feedback import (
    AnalyzeFeedbackRequest,
    AnalysisResponse,
    FeedbackSummary,
    FeedbackSegment,
    FeedbackDatasetResponse,
    StoredSegment,
    RenameDatasetRequest,
    AnalyticsReport,
)
from app.services.feedback_structuring_agent import FeedbackStructuringAgent
from app.services.video_optimization_agent import VideoOptimizationAgent
from app.services.analytics_agent import AnalyticsAgent
from app.services.feedback_dataset_service import FeedbackDatasetService
from app.services.project_service import ProjectService
from app.db.database import get_db
from app.utils.sensecap_export import (
    SENSECAP_CS_COLUMNS,
    normalise_sentiment_label  as _normalise_sentiment_label,
    sentiment_score            as _sentiment_score,
    safe_filename              as _safe_filename,
    build_sensecap_csv,
)

router = APIRouter()

_structuring_agent  = FeedbackStructuringAgent()
_optimization_agent = VideoOptimizationAgent()
_analytics_agent    = AnalyticsAgent()
_dataset_service    = FeedbackDatasetService()
_project_service    = ProjectService()

# Accepted file extensions for structured dataset upload
_ALLOWED_EXTENSIONS = {".json", ".csv", ".txt"}


# ── File upload endpoint ──────────────────────────────────────────────────────

@router.post("/upload-feedback", response_model=AnalysisResponse, status_code=201)
async def upload_feedback_file(
    project_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Accept a structured feedback file (.json or .csv) for a project.

    Expected JSON format:
        [
          {
            "timestamp": "00:34",       // MM:SS string or null
            "topic": "Camera",
            "sentiment": "Positive",    // Positive|Negative|Neutral|Suggestion|Complaint|Praise
            "summary": "Great shot",
            "confidence": 0.92
          },
          ...
        ]

    Expected CSV format (header row required):
        timestamp,topic,sentiment,summary,confidence
        00:34,Camera,Positive,Great shot,0.92
        ,Music,Negative,Too loud,0.78

    Business logic (parsing) lives in _parse_file().
    This route only handles HTTP concerns.
    """
    project = _project_service.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    filename = file.filename or ""
    ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in _ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Allowed: .json, .csv, .txt",
        )

    raw_bytes = await file.read()
    raw_text  = raw_bytes.decode("utf-8", errors="replace")

    # ── Stage 1: Parse file into FeedbackSegment list ─────────────────────────
    # .txt → FeedbackStructuringAgent (Gemini 2.5 Pro or regex fallback)
    # .json / .csv → structured parser
    try:
        if ext == ".txt":
            segments = _structuring_agent.parse(raw_text)  # HuggingFace zero-shot or regex fallback
            source   = "file_upload_txt"
        else:
            segments = _parse_file(raw_text, ext)
            source   = "file_upload"
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    if not segments:
        raise HTTPException(status_code=422, detail="File contained no valid feedback segments.")

    # ── Stage 2: Persist to DB BEFORE Agent 2 runs ───────────────────────────
    dataset = _dataset_service.save_dataset(
        db=db,
        project_id=project_id,
        raw_text=raw_text,
        segments=segments,
        source=source,
    )

    # ── Stage 3: Video Optimization Agent ────────────────────────────────────
    # Run scene detection + transcription in parallel so the agent gets
    # scene-anchored recommendations and speech-density analysis.
    scene_boundaries, transcript = _get_video_context(project)
    recommendations, editing_plan = _optimization_agent.analyze(
        segments=segments,
        video_metadata=project,
        scene_boundaries=scene_boundaries,
        transcript=transcript,
    )

    positive = sum(1 for s in segments if s.sentiment in ("Positive", "Praise"))
    negative = sum(1 for s in segments if s.sentiment in ("Negative", "Complaint"))
    neutral  = sum(1 for s in segments if s.sentiment in ("Neutral", "Question", "Suggestion"))

    return AnalysisResponse(
        dataset_id=dataset.id,
        feedback_summary=FeedbackSummary(
            positive=positive,
            negative=negative,
            neutral=neutral,
        ),
        timeline_insights=segments,
        optimization_recommendations=recommendations,
        editing_plan=editing_plan,
    )


# ── Text submit endpoint (kept for future LLM integration) ───────────────────

@router.post("/analyze-feedback", response_model=AnalysisResponse)
def analyze_feedback(
    body: AnalyzeFeedbackRequest,
    db: Session = Depends(get_db),
):
    """
    Submit raw unstructured feedback text.
    Uses HuggingFace zero-shot classification (facebook/bart-large-mnli) or regex fallback.
    """
    project = _project_service.get_project(body.project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not body.feedback.strip():
        raise HTTPException(status_code=400, detail="Feedback text must not be empty")

    segments = _structuring_agent.parse(body.feedback)

    dataset = _dataset_service.save_dataset(
        db=db,
        project_id=body.project_id,
        raw_text=body.feedback,
        segments=segments,
        source="manual_paste",
    )

    scene_boundaries, transcript = _get_video_context(project)
    recommendations, editing_plan = _optimization_agent.analyze(
        segments=segments,
        video_metadata=project,
        scene_boundaries=scene_boundaries,
        transcript=transcript,
    )

    positive = sum(1 for s in segments if s.sentiment in ("Positive", "Praise"))
    negative = sum(1 for s in segments if s.sentiment in ("Negative", "Complaint"))
    neutral  = sum(1 for s in segments if s.sentiment in ("Neutral", "Question", "Suggestion"))

    return AnalysisResponse(
        dataset_id=dataset.id,
        feedback_summary=FeedbackSummary(positive=positive, negative=negative, neutral=neutral),
        timeline_insights=segments,
        optimization_recommendations=recommendations,
        editing_plan=editing_plan,
    )


# ── Dataset read / delete endpoints ──────────────────────────────────────────

@router.get(
    "/feedback-datasets/{project_id}",
    response_model=list[FeedbackDatasetResponse],
)
def list_datasets(project_id: str, db: Session = Depends(get_db)):
    """Return all stored feedback datasets for a project, newest first."""
    if not _project_service.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return [_serialise(ds) for ds in _dataset_service.get_datasets_for_project(db, project_id)]


@router.get("/feedback-dataset/{dataset_id}", response_model=FeedbackDatasetResponse)
def get_dataset(dataset_id: str, db: Session = Depends(get_db)):
    """Return a single stored dataset with all its segments."""
    ds = _dataset_service.get_dataset_by_id(db, dataset_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset not found")
    return _serialise(ds)


@router.patch("/feedback-dataset/{dataset_id}/rename", response_model=FeedbackDatasetResponse)
def rename_dataset(dataset_id: str, body: RenameDatasetRequest, db: Session = Depends(get_db)):
    """Set a user-defined name on a stored dataset."""
    ds = _dataset_service.rename_dataset(db, dataset_id, body.name)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset not found")
    return _serialise(ds)


@router.get("/analytics/{dataset_id}", response_model=AnalyticsReport)
def get_analytics(dataset_id: str, db: Session = Depends(get_db)):
    """
    Run the Analytics Agent on a saved dataset.
    Result is cached in the DB after first computation — subsequent calls
    return the cached report instantly without recomputing.
    """
    ds = _dataset_service.get_dataset_by_id(db, dataset_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset not found")

    # Return cached report if available
    cached = _dataset_service.get_analytics_cache(db, dataset_id)
    if cached:
        return AnalyticsReport(**cached)

    # Compute fresh and cache
    segments = [
        FeedbackSegment(
            timestamp=seg.timestamp,
            topic=seg.topic,
            sentiment=seg.sentiment,
            summary=seg.summary,
            confidence=seg.confidence,
        )
        for seg in ds.segments
    ]
    report = _analytics_agent.analyze(segments)
    _dataset_service.set_analytics_cache(db, dataset_id, report.model_dump_json())
    return report


@router.get("/export-dataset/{dataset_id}/csv")
def export_dataset_csv(dataset_id: str, db: Session = Depends(get_db)):
    """
    Export a persisted ClipSense feedback dataset as a Sensecap-compatible CSV.

    The CSV uses SENSECAP_CS_COLUMNS as its canonical schema (defined in
    app/utils/sensecap_export.py). Only fields that ClipSense genuinely
    possesses are included. Geography, engagement, and ROI columns are
    intentionally absent so that Sensecap renders honest "unavailable" states
    rather than fabricated zeros.

    Security posture:
    - No authentication exists in this application; this endpoint inherits
      that posture. Dataset IDs are UUIDs — enumeration is impractical.
    - The filename is sanitized before use in Content-Disposition.
    - Text fields are written via csv.DictWriter which handles quoting,
      commas, newlines, and Unicode correctly.
    - CSV injection: fields beginning with =, +, -, @ are not prefixed with
      a tab because this CSV is consumed by Pandas (not opened in Excel by
      end users directly). The risk is documented and accepted.

    Performance:
    - Datasets are loaded fully into memory. Acceptable for the current POC
      where datasets are small (typically < 10 000 segments). For large
      production datasets, replace build_sensecap_csv with a streaming
      generator and return a StreamingResponse with a generator body.
    """
    ds = _dataset_service.get_dataset_by_id(db, dataset_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset not found")

    if not ds.segments:
        raise HTTPException(
            status_code=422,
            detail="Dataset exists but contains no segments and cannot be exported.",
        )

    dataset_label = ds.name or ""
    csv_bytes = build_sensecap_csv(ds.segments, dataset_label)
    safe_name = _safe_filename(dataset_label, dataset_id[:8])
    filename  = f"sensecap_{safe_name}.csv"

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export-dataset/{dataset_id}")
def export_dataset_excel(dataset_id: str, db: Session = Depends(get_db)):
    """Export a dataset as .xlsx — segments sheet + raw text sheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ds = _dataset_service.get_dataset_by_id(db, dataset_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset not found")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Segments ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Segments"
    headers = ["#", "Timestamp", "Topic", "Sentiment", "Summary", "Confidence", "Created At"]
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="94A3B8")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sentiment_colors = {
        "Positive": "166534", "Praise": "166534",
        "Negative": "991B1B", "Complaint": "991B1B",
        "Suggestion": "92400E", "Question": "1E3A5F",
        "Neutral": "374151",
    }

    for seg in ds.segments:
        row = [
            seg.position + 1,
            seg.timestamp or "—",
            seg.topic,
            seg.sentiment,
            seg.summary,
            round(seg.confidence * 100, 1),
            seg.created_at.strftime("%Y-%m-%d %H:%M"),
        ]
        ws.append(row)
        color = sentiment_colors.get(seg.sentiment, "374151")
        ws.cell(row=ws.max_row, column=4).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=ws.max_row, column=4).font = Font(color="FFFFFF", bold=True)

    ws.column_dimensions["E"].width = 60
    for col in ["A","B","C","D","F","G"]:
        ws.column_dimensions[col].width = 18

    # ── Sheet 2: Raw Feedback ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Feedback")
    ws2.append(["Dataset ID", ds.id])
    ws2.append(["Project ID", ds.project_id])
    ws2.append(["Source",     ds.source])
    ws2.append(["Name",       ds.name or ""])
    ws2.append(["Created At", ds.created_at.strftime("%Y-%m-%d %H:%M:%S")])
    ws2.append([])
    ws2.append(["Raw Text"])
    ws2.cell(row=7, column=1).font = Font(bold=True)
    for line in (ds.raw_text or "").splitlines():
        ws2.append([line])
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    label = (ds.name or dataset_id[:8]).replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="clipsense_{label}.xlsx"'},
    )


@router.delete("/feedback-dataset/{dataset_id}", status_code=204)
def delete_dataset(dataset_id: str, db: Session = Depends(get_db)):
    """Delete a stored dataset and all its segments."""
    if not _dataset_service.delete_dataset(db, dataset_id):
        raise HTTPException(status_code=404, detail="Dataset not found")


# ── Private helpers ───────────────────────────────────────────────────────────

def _get_video_context(project: dict) -> tuple[list, dict]:
    """
    Run scene detection and Whisper transcription in parallel for the project
    video. Returns (scene_boundaries, transcript). Both default to empty
    structures on failure so the caller never needs to handle None.
    """
    import os
    from concurrent.futures import ThreadPoolExecutor
    from app.utils.scene_detector import detect_scenes
    from app.utils.transcript import transcribe
    from app.utils.storage import UPLOAD_DIR

    video_path = project.get("file_path", "")
    if not video_path or not os.path.exists(video_path):
        pid = project.get("id", "")
        for ext in (".mp4", ".mov", ".avi", ".mkv", ".webm"):
            candidate = os.path.join(UPLOAD_DIR, f"{pid}{ext}")
            if os.path.exists(candidate):
                video_path = candidate
                break

    if not video_path or not os.path.exists(video_path):
        return [], {"segments": [], "words": [], "language": "", "full_text": ""}

    try:
        with ThreadPoolExecutor(max_workers=2) as pool:
            fut_scenes     = pool.submit(detect_scenes, video_path)
            fut_transcript = pool.submit(transcribe, video_path)
            return fut_scenes.result(), fut_transcript.result()
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("feedback: video context extraction failed: %s", exc)
        return [], {"segments": [], "words": [], "language": "", "full_text": ""}


def _parse_file(raw_text: str, ext: str) -> list[FeedbackSegment]:
    """
    Parse a structured feedback file into FeedbackSegment objects.

    Supports .json and .csv.
    Raises ValueError with a descriptive message on invalid structure.

    Structured files (.json, .csv) are parsed directly.
    Unstructured .txt files go through FeedbackStructuringAgent (HuggingFace or regex).
    """
    if ext == ".json":
        return _parse_json(raw_text)
    if ext == ".csv":
        return _parse_csv(raw_text)
    raise ValueError(f"Unsupported extension: {ext}")


def _parse_json(raw_text: str) -> list[FeedbackSegment]:
    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON: {exc}") from exc

    if not isinstance(data, list):
        raise ValueError("JSON file must contain a top-level array of segment objects.")

    segments: list[FeedbackSegment] = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"Item at index {i} is not an object.")
        try:
            segments.append(FeedbackSegment(
                timestamp=item.get("timestamp"),
                topic=item.get("topic", "General"),
                sentiment=item.get("sentiment", "Neutral"),
                summary=str(item.get("summary", item.get("comment", ""))),
                confidence=float(item.get("confidence", 0.75)),
            ))
        except Exception as exc:
            raise ValueError(f"Invalid segment at index {i}: {exc}") from exc

    return segments


def _parse_csv(raw_text: str) -> list[FeedbackSegment]:
    reader = csv.DictReader(io.StringIO(raw_text))

    required = {"sentiment", "summary"}
    if reader.fieldnames:
        missing = required - {f.strip().lower() for f in reader.fieldnames if f}
        if missing:
            raise ValueError(
                f"CSV is missing required columns: {', '.join(missing)}. "
                f"Required: timestamp, topic, sentiment, summary, confidence"
            )

    segments: list[FeedbackSegment] = []
    for i, row in enumerate(reader):
        # Normalise keys to lowercase and strip whitespace
        row = {k.strip().lower(): (v.strip() if v else "") for k, v in row.items()}
        try:
            segments.append(FeedbackSegment(
                timestamp=row.get("timestamp") or None,
                topic=row.get("topic", "General") or "General",
                sentiment=row.get("sentiment", "Neutral") or "Neutral",
                summary=row.get("summary", row.get("comment", "")),
                confidence=float(row.get("confidence", 0.75) or 0.75),
            ))
        except Exception as exc:
            raise ValueError(f"Invalid row at line {i + 2}: {exc}") from exc

    return segments


def _serialise(ds) -> FeedbackDatasetResponse:
    return FeedbackDatasetResponse(
        id=ds.id,
        project_id=ds.project_id,
        name=ds.name,
        source=ds.source,
        created_at=ds.created_at.isoformat(),
        segment_count=len(ds.segments),
        segments=[
            StoredSegment(
                id=seg.id,
                position=seg.position,
                timestamp=seg.timestamp,
                topic=seg.topic,
                sentiment=seg.sentiment,
                summary=seg.summary,
                confidence=seg.confidence,
                created_at=seg.created_at.isoformat(),
            )
            for seg in ds.segments
        ],
    )
